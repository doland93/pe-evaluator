import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os, re, threading
from collections import defaultdict
from difflib import SequenceMatcher

import pdfplumber, fitz
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════
#  명단 파서 — 절대 기준
# ═══════════════════════════════════════════════

def parse_roster(path):
    """
    반환: {
      "1": [{"display":"김민규","korean":"김민규","english":None}, ...],
      "2": [{"display":"닐루파르","korean":"닐루파르","english":"NILUFAR HUSEYNLI"}, ...]
    }
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [[str(c).strip() if c is not None else "" for c in row]
            for row in ws.iter_rows(values_only=True)]

    # 헤더 행 탐지
    header_idx = team_col = name_col = None
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            cl = cell.lower().strip()
            if cl in ('팀','team','조'): team_col = j
            if re.search(r'이름|name|학생명|성명', cl): name_col = j
        if team_col is not None and name_col is not None:
            header_idx = i; break
    # 헤더 못 찾으면 첫 숫자 행 기준
    if header_idx is None:
        for i, row in enumerate(rows):
            if row and re.match(r'^\d+$', row[0]):
                header_idx = i-1; team_col = 0; name_col = 1; break
    if header_idx is None: return {}

    roster = defaultdict(list)
    for row in rows[header_idx+1:]:
        if not any(row): continue
        team_val = row[team_col] if team_col < len(row) else ""
        name_val = row[name_col] if name_col < len(row) else ""
        tm = re.search(r'(\d+)', team_val)
        if not tm: continue
        team = tm.group(1)

        # 외국인: 줄바꿈+괄호 안 한글
        m = re.search(r'[\n\r]*[（(]([가-힣]{2,})[）)]', name_val)
        if m:
            korean = m.group(1).strip()
            english = re.sub(r'[\n\r]*[（(].*', '', name_val).strip().upper()
            roster[team].append({"display": korean, "korean": korean, "english": english})
        else:
            korean = "".join(re.findall(r'[가-힣]+', name_val))
            if len(korean) >= 2:
                roster[team].append({"display": korean, "korean": korean, "english": None})
    return dict(roster)


def find_in_roster(raw_name, roster):
    """
    raw_name(PDF에서 파싱된 이름)을 명단에서 찾아
    (team, display_name) 반환. 없으면 (None, None).
    
    매칭 우선순위:
    1) 한글 완전일치
    2) 영문 첫 단어(이름) 일치
    3) 영문 전체 포함
    4) 괄호 안 영문으로 매칭
    5) 유사도 매칭
    """
    if not raw_name: return None, None
    raw = raw_name.strip().replace('\n',' ').replace('\r','')

    # raw에서 한글/영문 추출
    m = re.search(r'[（(]([가-힣]{2,})[）)]', raw)
    raw_kor = m.group(1) if m else "".join(re.findall(r'[가-힣]+', raw))
    
    m2 = re.search(r'[（(]([a-zA-Z]{3,})[）)]', raw)
    raw_eng_paren = m2.group(1).upper() if m2 else None
    
    # 영문 이름에서 첫 단어(이름 부분) 추출
    raw_words = re.findall(r'[A-Za-z]+', raw)
    raw_first_word = raw_words[0].upper() if raw_words else None
    raw_eng_full = re.sub(r'\s+','', raw.upper())
    raw_eng_nospace = "".join(raw_words).upper() if raw_words else ""

    best_team = best_disp = None
    best_score = 0

    for team, members in roster.items():
        for mb in members:
            # 1) 한글 완전일치
            if raw_kor and raw_kor == mb["korean"]:
                return team, mb["display"]
            
            if mb["english"]:
                eng_full = re.sub(r'\s+','', mb["english"])
                eng_words = mb["english"].split()
                eng_first = eng_words[0].upper() if eng_words else ""
                
                # 2) 영문 첫 단어(이름) 완전일치
                if raw_first_word and raw_first_word == eng_first:
                    return team, mb["display"]
                
                # 3) 영문 전체 포함
                if raw_eng_nospace and (raw_eng_nospace == eng_full or 
                                         raw_eng_nospace in eng_full or 
                                         eng_full in raw_eng_nospace):
                    return team, mb["display"]
                
                # 4) 괄호 안 영문으로 매칭
                if raw_eng_paren and eng_first.startswith(raw_eng_paren):
                    return team, mb["display"]
                
                # 영문 유사도 (첫 단어 기준)
                if raw_first_word and eng_first:
                    s = SequenceMatcher(None, raw_first_word, eng_first).ratio()
                    if s > best_score:
                        best_score = s; best_team = team; best_disp = mb["display"]
                
                # 영문 전체 유사도
                s = SequenceMatcher(None, raw_eng_full, eng_full).ratio()
                if s > best_score:
                    best_score = s; best_team = team; best_disp = mb["display"]
            
            # 5) 한글 유사도
            if raw_kor and mb["korean"]:
                s = SequenceMatcher(None, raw_kor, mb["korean"]).ratio()
                if s > best_score:
                    best_score = s; best_team = team; best_disp = mb["display"]

    if best_score >= 0.5:
        return best_team, best_disp
    return None, None


# ═══════════════════════════════════════════════
#  PDF / DOCX 파서 — 점수만 추출
# ═══════════════════════════════════════════════

def _pdf_text(path):
    text = ""
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t: text += t + "\n"
    except: pass
    if not text.strip():
        try:
            doc = fitz.open(path)
            for page in doc: text += page.get_text() + "\n"
        except: pass
    return text

def _parse_docx(path):
    doc = Document(path)
    result = {"evaluator_raw": None, "scores_raw": []}
    for table in doc.tables:
        for row in table.rows:
            cells = [c.text.strip() for c in row.cells]
            if len(cells) < 2: continue
            if re.match(r'Name\s*:', cells[0], re.I):
                nm = re.sub(r'Name\s*:\s*', '', cells[0], flags=re.I).strip()
                if nm: result["evaluator_raw"] = nm
            if re.match(r'^[1-5]\s*[.。]', cells[0]):
                name_raw = re.sub(r'^[1-5]\s*[.。]\s*', '', cells[0]).strip()
                name_raw = re.sub(r'\d{7,10}', '', name_raw).strip()
                pts_raw = cells[-1].strip()
                if name_raw and re.match(r'^\d+$', pts_raw):
                    result["scores_raw"].append((name_raw, int(pts_raw)))
    return result

def _parse_pdf(path):
    text = _pdf_text(path)
    result = {"evaluator_raw": None, "scores_raw": []}
    m = re.search(r"Name\s*[:：]\s*([가-힣]+(?:\s+[가-힣]+)?)", text, re.I)
    if m: result["evaluator_raw"] = m.group(1).strip()
    if not result["evaluator_raw"]:
        m = re.search(r"Name\s*[:：]\s*([A-Za-z\s\(\)가-힣]+?)(?:\n|Team)", text, re.I)
        if m: result["evaluator_raw"] = m.group(1).strip()
    for line in text.split("\n"):
        line = line.strip()
        m = re.match(
            r"^([1-5])\s*[.。]\s*(?:\d{7,10}\s+)?([가-힣A-Za-z\(\)]{2,20})\s+(\d{1,3})\s*$",
            line)
        if m: result["scores_raw"].append((m.group(2).strip(), int(m.group(3))))
    return result

def parse_file(path):
    try:
        if path.lower().endswith('.docx'):
            return _parse_docx(path)
        else:
            return _parse_pdf(path)
    except Exception as e:
        return {"evaluator_raw": None, "scores_raw": [], "_error": str(e)}


# ═══════════════════════════════════════════════
#  이슈 감지
# ═══════════════════════════════════════════════

def detect_issues(records):
    issues = []
    for rec in records:
        rec_issues = []
        total = sum(p for _, p in rec.get("scores", []))
        if rec.get("scores"):
            if total == 0:
                rec_issues.append("점수 합 = 0")
            elif total != 100:
                if all(p <= 100 for _, p in rec["scores"]) and total > 100:
                    rec_issues.append(f"⚠ 개인별 100점 방식 의심 (합={total})")
                else:
                    rec_issues.append(f"점수 합 = {total}")
        if rec.get("_no_match"):
            rec_issues.append("⚠ 명단에서 평가자 이름 못 찾음")
        issues.append(rec_issues)
    return issues


# ═══════════════════════════════════════════════
#  엑셀 생성
# ═══════════════════════════════════════════════

_NAVY="1F4E79"; _LBLUE="D6E4F0"; _LBLUE2="EBF5FB"
_GREEN="E8F5E9"; _WARN="FFF9C4"; _ERR="FFCDD2"; _WHITE="FFFFFF"

def _fill(c): return PatternFill("solid", fgColor=c)
def _side(w="thin", c="BBBBBB"): return Side(style=w, color=c)
def _hdr(ws, r, col, val):
    c = ws.cell(row=r, column=col, value=val)
    c.font = Font(bold=True, color="FFFFFF", name="Malgun Gothic", size=10)
    c.fill = _fill(_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=_side(), right=_side(), top=_side(), bottom=_side())

def _cell(ws, r, col, val, fill=None, bold=False, fmt=None):
    c = ws.cell(row=r, column=col, value=val)
    c.font = Font(bold=bold, name="Malgun Gothic", size=10)
    if fill: c.fill = _fill(fill)
    c.alignment = Alignment(horizontal="center")
    c.border = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
    if fmt: c.number_format = fmt
    return c

def build_excel(records, out_path, roster):
    # ── 팀/멤버 구성: 명단 기준 ──
    if roster:
        team_order = {t: [mb["display"] for mb in members]
                      for t, members in roster.items()}
    else:
        team_members = defaultdict(set)
        for rec in records:
            t = rec.get("team","?")
            if rec.get("evaluator"): team_members[t].add(rec["evaluator"])
            for name, _ in rec.get("scores",[]): team_members[t].add(name)
        team_order = {t: sorted(m) for t, m in team_members.items()}

    # eval_map[(evaluator, recipient)] = pts
    eval_map = {}
    for rec in records:
        ev = rec.get("evaluator")
        for name, pts in rec.get("scores", []):
            if ev: eval_map[(ev, name)] = pts

    teams = sorted(team_order.keys(), key=lambda x: int(x) if x.isdigit() else x)
    student_order = [(t, n) for t in teams for n in team_order[t]]
    max_members = max(len(m) for m in team_order.values())

    wb = Workbook()
    ws = wb.active; ws.title = "PE"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C2"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 7
    for i in range(max_members):
        ws.column_dimensions[get_column_letter(3+i)].width = 9
    ws.column_dimensions[get_column_letter(3+max_members)].width = 7
    ws.column_dimensions[get_column_letter(4+max_members)].width = 9

    hdrs = ["학생명","Group"] + [str(i+1) for i in range(max_members)] + ["합","Weight"]
    for col, h in enumerate(hdrs, 1): _hdr(ws, 1, col, h)
    ws.row_dimensions[1].height = 22

    team_rows = {}; row = 2
    for (team, name) in student_order:
        if team not in team_rows: team_rows[team] = {"start": row}
        team_rows[team]["end"] = row
        members = team_order[team]
        alt = int(team) % 2 == 0 if team.isdigit() else False
        bf = _LBLUE2 if alt else _LBLUE

        _cell(ws, row, 1, name, fill=bf)
        _cell(ws, row, 2, int(team) if team.isdigit() else team, fill=bf)

        # 각 열 = 팀원 순번, 본인=회색, 나머지=그 팀원이 name에게 준 점수
        for i, evaluator in enumerate(members):
            col = 3 + i
            if evaluator == name:
                _cell(ws, row, col, None, fill="D0D0D0")
            else:
                pts = eval_map.get((evaluator, name), None)
                fill = _WARN if pts is None else _WHITE
                _cell(ws, row, col, pts, fill=fill)

        for i in range(len(members), max_members):
            _cell(ws, row, 3+i, None, fill="F0F0F0")

        # 합 = SUM 수식
        sc = 3+max_members
        sum_range = f"{get_column_letter(3)}{row}:{get_column_letter(2+max_members)}{row}"
        c = ws.cell(row=row, column=sc, value=f"=SUM({sum_range})")
        c.font = Font(bold=True, name="Malgun Gothic", size=10)
        c.fill = _fill(_GREEN); c.alignment = Alignment(horizontal="center")
        c.border = Border(left=_side(), right=_side(), top=_side(), bottom=_side())

        # Weight = 합/100 수식
        wc = 4+max_members
        c = ws.cell(row=row, column=wc, value=f"={get_column_letter(sc)}{row}/100")
        c.font = Font(bold=True, name="Malgun Gothic", size=10)
        c.fill = _fill(_GREEN); c.alignment = Alignment(horizontal="center")
        c.number_format = "0.00"
        c.border = Border(left=_side(), right=_side(), top=_side(), bottom=_side())

        ws.row_dimensions[row].height = 18; row += 1

    # 팀 구분선 + Weight 합계
    total_cols = 4+max_members
    for team, rng in team_rows.items():
        s, e = rng["start"], rng["end"]
        members = team_order[team]
        for r in range(s, e+1):
            for col in range(1, total_cols+1):
                cell = ws.cell(row=r, column=col)
                old = cell.border
                cell.border = Border(left=old.left, right=old.right,
                    top=_side("medium",_NAVY) if r==s else old.top,
                    bottom=_side("medium",_NAVY) if r==e else old.bottom)
        # 팀 weight 합계 = SUM 수식
        wc = 4+max_members
        weight_range = f"{get_column_letter(wc)}{s}:{get_column_letter(wc)}{e}"
        c = ws.cell(row=e, column=total_cols+1, value=f"=SUM({weight_range})")
        c.font = Font(italic=True, color="888888", name="Malgun Gothic", size=9)
        c.number_format = "0.00"
        c.alignment = Alignment(horizontal="center")

    # ── 제출현황 시트 ──
    ws2 = wb.create_sheet("제출현황")
    ws2.sheet_view.showGridLines = False
    for col, (h, w) in enumerate(zip(
            ["이름","팀","제출여부","부여점수합","비고"],
            [12, 8, 14, 12, 20]), 1):
        _hdr(ws2, 1, col, h)
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 22

    submitted = {rec["evaluator"] for rec in records if rec.get("evaluator")}
    for i, (team, name) in enumerate(student_order, 2):
        ok = name in submitted
        pts_given = sum(p for (ev,_),p in eval_map.items() if ev == name)
        note = ""
        if ok and pts_given != 100:
            note = f"⚠ 합계 {pts_given}점"
        elif not ok:
            note = "미제출"
        rf = _WHITE if ok else _ERR
        for col, v in enumerate([name, team, "✔ 제출" if ok else "✘ 미제출",
                                  pts_given if ok else "-", note], 1):
            c2 = ws2.cell(row=i, column=col, value=v)
            c2.font = Font(name="Malgun Gothic", size=10,
                           color="006400" if ok else "C62828")
            c2.fill = _fill(rf)
            c2.alignment = Alignment(horizontal="center")
            c2.border = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
        ws2.row_dimensions[i].height = 17

    wb.save(out_path)


# ═══════════════════════════════════════════════
#  GUI — 메인 창
# ═══════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PE 집계기  —  Peer Evaluation Processor")
        self.geometry("740x660")
        self.minsize(620, 560)
        self.configure(bg="#F0F4F8")
        self.files = []
        self.roster_path = None
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=f"#{_NAVY}", pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📋  Peer Evaluation 자동 집계기",
                 font=("Malgun Gothic",14,"bold"), fg="white", bg=f"#{_NAVY}").pack()
        tk.Label(hdr, text="① 명단 업로드 → ② PDF/DOCX 추가 → ③ 파싱 확인 → ④ 엑셀 생성",
                 font=("Malgun Gothic",9), fg="#BDD7EE", bg=f"#{_NAVY}").pack()

        # 명단
        rf = tk.LabelFrame(self, text=" 👥 팀 명단 엑셀 (권장) ",
                           font=("Malgun Gothic",10,"bold"),
                           bg="#F0F4F8", fg=f"#{_NAVY}", bd=2, relief="groove", padx=8, pady=6)
        rf.pack(fill="x", padx=16, pady=(14,4))
        ri = tk.Frame(rf, bg="#F0F4F8"); ri.pack(fill="x")
        self.roster_var = tk.StringVar(value="명단 파일 없음")
        tk.Label(ri, textvariable=self.roster_var, font=("Consolas",9),
                 bg="white", fg="#555", relief="solid", bd=1,
                 anchor="w", padx=6).pack(side="left", fill="x", expand=True, padx=(0,6))
        tk.Button(ri, text="명단 선택", bg="#8E44AD", fg="white",
                  font=("Malgun Gothic",9,"bold"), relief="flat",
                  padx=10, pady=4, cursor="hand2",
                  command=self._load_roster).pack(side="right")

        # 파일 목록
        lf = tk.LabelFrame(self, text=" 📂 PE 제출 파일 (PDF / DOCX) ",
                           font=("Malgun Gothic",10,"bold"),
                           bg="#F0F4F8", fg=f"#{_NAVY}", bd=2, relief="groove", padx=8, pady=6)
        lf.pack(fill="both", expand=True, padx=16, pady=(4,4))
        self.lb = tk.Listbox(lf, font=("Consolas",9), selectmode="extended",
                             height=9, bg="white", fg="#222",
                             selectbackground=f"#{_NAVY}", selectforeground="white",
                             bd=0, highlightthickness=1, highlightbackground="#CCC")
        sb = ttk.Scrollbar(lf, orient="vertical", command=self.lb.yview)
        self.lb.configure(yscrollcommand=sb.set)
        self.lb.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        bf = tk.Frame(self, bg="#F0F4F8"); bf.pack(fill="x", padx=16, pady=4)
        S = dict(font=("Malgun Gothic",10,"bold"), relief="flat", padx=12, pady=6, cursor="hand2")
        tk.Button(bf, text="+ 파일 추가", bg="#2E86C1", fg="white",
                  command=self._add, **S).pack(side="left", padx=(0,6))
        tk.Button(bf, text="🗑 선택 삭제", bg="#7F8C8D", fg="white",
                  command=self._remove, **S).pack(side="left", padx=(0,6))
        tk.Button(bf, text="전체 초기화", bg="#C0392B", fg="white",
                  command=self._clear, **S).pack(side="left")

        # 저장 위치
        pf = tk.LabelFrame(self, text=" 💾 저장 위치 ",
                           font=("Malgun Gothic",10,"bold"),
                           bg="#F0F4F8", fg=f"#{_NAVY}", bd=2, relief="groove", padx=8, pady=6)
        pf.pack(fill="x", padx=16, pady=(6,4))
        pi = tk.Frame(pf, bg="#F0F4F8"); pi.pack(fill="x")
        self.path_var = tk.StringVar(
            value=os.path.join(os.path.expanduser("~"), "Desktop", "PE_결과.xlsx"))
        tk.Entry(pi, textvariable=self.path_var, font=("Consolas",9),
                 bg="white", relief="solid", bd=1).pack(side="left", fill="x",
                                                         expand=True, padx=(0,6))
        tk.Button(pi, text="찾기", bg="#5D6D7E", fg="white",
                  font=("Malgun Gothic",9,"bold"), relief="flat",
                  padx=10, pady=4, cursor="hand2", command=self._browse).pack(side="right")

        self.run_btn = tk.Button(self, text="🔍  파싱 & 확인",
                                 bg=f"#{_NAVY}", fg="white",
                                 font=("Malgun Gothic",13,"bold"),
                                 relief="flat", padx=24, pady=10,
                                 cursor="hand2", command=self._run)
        self.run_btn.pack(pady=(10,4))

        self.status = tk.StringVar(value="명단 파일과 PE 파일을 추가하고 '파싱 & 확인' 버튼을 눌러주세요.")
        tk.Label(self, textvariable=self.status, font=("Malgun Gothic",9),
                 bg="#DDE3EA", fg="#333", anchor="w", padx=10, pady=5
                 ).pack(fill="x", side="bottom")

    def _load_roster(self):
        p = filedialog.askopenfilename(
            title="팀 명단 엑셀 선택",
            filetypes=[("Excel","*.xlsx *.xls")])
        if p:
            self.roster_path = p
            self.roster_var.set(f"✔ {os.path.basename(p)}")
            self.status.set(f"명단 로드됨: {os.path.basename(p)}")

    def _add(self):
        paths = filedialog.askopenfilenames(
            title="PDF 또는 DOCX 파일 선택",
            filetypes=[("지원 파일","*.pdf *.docx"),("PDF","*.pdf"),("Word","*.docx")])
        for p in paths:
            if p not in self.files:
                self.files.append(p)
                self.lb.insert("end", f"  {os.path.basename(p)}")
        self.status.set(f"총 {len(self.files)}개 파일 로드됨")

    def _remove(self):
        for i in reversed(self.lb.curselection()):
            self.lb.delete(i); self.files.pop(i)
        self.status.set(f"총 {len(self.files)}개 파일")

    def _clear(self):
        self.files.clear(); self.lb.delete(0, "end")
        self.status.set("초기화되었습니다.")

    def _browse(self):
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")], initialfile="PE_결과.xlsx")
        if p: self.path_var.set(p)

    def _run(self):
        if not self.files and not self.roster_path:
            messagebox.showwarning("파일 없음", "파일을 먼저 추가해주세요.")
            return
        self.run_btn.config(state="disabled", text="⏳  파싱 중...")
        threading.Thread(target=self._parse_all, daemon=True).start()

    def _parse_all(self):
        # 1) 명단 파싱
        roster = None
        if self.roster_path:
            try:
                roster = parse_roster(self.roster_path)
                total = sum(len(v) for v in roster.values())
                self.after(0, lambda: self.status.set(f"명단 파싱 완료: {total}명"))
            except Exception as e:
                self.after(0, lambda: self.status.set(f"명단 파싱 오류: {e}"))

        # 2) PDF/DOCX 파싱 — 점수만 추출, 이름은 명단 기준으로 매칭
        records = []
        errors = []
        for path in self.files:
            fname = os.path.basename(path)
            try:
                raw = parse_file(path)
                rec = {"evaluator": None, "team": None, "scores": [], "_file": fname}

                # 평가자 이름 → 명단 매칭
                ev_raw = raw.get("evaluator_raw","")
                if roster and ev_raw:
                    team, display = find_in_roster(ev_raw, roster)
                    rec["evaluator"] = display
                    rec["team"] = team
                    if not display:
                        rec["_no_match"] = True
                        rec["evaluator"] = ev_raw  # 매칭 실패 시 원본 유지
                else:
                    rec["evaluator"] = ev_raw

                # 점수 목록 — 피평가자 이름도 명단 매칭
                for name_raw, pts in raw.get("scores_raw", []):
                    if roster:
                        _, disp = find_in_roster(name_raw, roster)
                        rec["scores"].append((disp if disp else name_raw, pts))
                    else:
                        rec["scores"].append((name_raw, pts))

                records.append(rec)
                self.after(0, lambda n=fname: self.status.set(f"파싱: {n}"))
            except Exception as e:
                errors.append(f"{fname}: {e}")

        if not records and not roster:
            self.after(0, lambda: messagebox.showerror("오류","파싱된 데이터가 없습니다."))
            self.after(0, lambda: self.run_btn.config(state="normal", text="🔍  파싱 & 확인"))
            return

        issues = detect_issues(records)
        self.after(0, lambda: self._show_review(records, issues, errors, self.path_var.get(), roster))
        self.after(0, lambda: self.run_btn.config(state="normal", text="🔍  파싱 & 확인"))

    def _show_review(self, records, issues, errors, out_path, roster=None):
        ReviewWindow(self, records, issues, errors, out_path, roster)


# ═══════════════════════════════════════════════
#  GUI — 검토/수정 창
# ═══════════════════════════════════════════════

class ReviewWindow(tk.Toplevel):
    def __init__(self, parent, records, issues, errors, out_path, roster=None):
        super().__init__(parent)
        self.title("파싱 결과 확인 및 수정")
        self.geometry("1060x680")
        self.minsize(800, 500)
        self.configure(bg="#F0F4F8")
        self.records = records
        self.issues  = issues
        self.errors  = errors
        self.out_path = out_path
        self.roster   = roster
        self._build()
        self._populate()

    def _build(self):
        top = tk.Frame(self, bg=f"#{_NAVY}", pady=10); top.pack(fill="x")
        tk.Label(top, text="📝  파싱 결과 확인  —  수정 후 엑셀 생성",
                 font=("Malgun Gothic",12,"bold"), fg="white", bg=f"#{_NAVY}").pack()
        tk.Label(top, text="팀 번호 더블클릭 = 수정 | 노란=경고 | 빨간=오류/미제출",
                 font=("Malgun Gothic",9), fg="#BDD7EE", bg=f"#{_NAVY}").pack()

        tf = tk.Frame(self, bg="#F0F4F8"); tf.pack(fill="both", expand=True, padx=14, pady=(10,4))
        cols = ("이름", "팀", "점수 목록", "합계", "이슈", "파일명")
        self.tree = ttk.Treeview(tf, columns=cols, show="headings", height=20)
        widths = [90, 60, 400, 60, 220, 180]
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="center" if w < 200 else "w")
        vsb = ttk.Scrollbar(tf, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(tf, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tf.rowconfigure(0, weight=1); tf.columnconfigure(0, weight=1)

        self.tree.tag_configure("ok",      background="#FFFFFF")
        self.tree.tag_configure("warn",    background="#FFF9C4")
        self.tree.tag_configure("err",     background="#FFCDD2")
        self.tree.tag_configure("missing", background="#FFE0E0")
        self.tree.bind("<Double-1>", self._on_double_click)

        # 범례
        leg = tk.Frame(self, bg="#F0F4F8"); leg.pack(fill="x", padx=14, pady=(0,4))
        for color, text in [
            ("#FFF9C4","개인별 100점 방식 의심"),
            ("#FFCDD2","오류 / 명단 매칭 실패"),
            ("#FFE0E0","미제출 (명단 기준)")]:
            tk.Frame(leg, bg=color, width=16, height=16, relief="solid", bd=1
                     ).pack(side="left", padx=(8,3), pady=2)
            tk.Label(leg, text=text, font=("Malgun Gothic",8),
                     bg="#F0F4F8", fg="#555").pack(side="left", padx=(0,14))

        bf = tk.Frame(self, bg="#F0F4F8"); bf.pack(fill="x", padx=14, pady=(4,10))
        self.summary_var = tk.StringVar()
        tk.Label(bf, textvariable=self.summary_var,
                 font=("Malgun Gothic",9), bg="#F0F4F8", fg="#555").pack(side="left")
        tk.Button(bf, text="⚡  엑셀 생성", bg="#1B5E20", fg="white",
                  font=("Malgun Gothic",12,"bold"), relief="flat",
                  padx=20, pady=8, cursor="hand2",
                  command=self._generate).pack(side="right")
        tk.Button(bf, text="닫기", bg="#7F8C8D", fg="white",
                  font=("Malgun Gothic",10,"bold"), relief="flat",
                  padx=14, pady=8, cursor="hand2",
                  command=self.destroy).pack(side="right", padx=(0,8))

    def _populate(self):
        self.tree.delete(*self.tree.get_children())
        submitted = {rec.get("evaluator") for rec in self.records if rec.get("evaluator")}
        warn_cnt = err_cnt = missing_cnt = ok_cnt = 0

        for i, rec in enumerate(self.records):
            ev    = rec.get("evaluator") or "?"
            team  = rec.get("team") or "?"
            scores_str = "  ".join(f"{n}:{p}" for n,p in rec.get("scores",[])) or "(없음)"
            total = sum(p for _,p in rec.get("scores",[]))
            iss   = self.issues[i] if i < len(self.issues) else []
            issue_str = " | ".join(iss) if iss else "✓ 정상"
            fname = rec.get("_file","")

            if "개인별 100점" in " ".join(iss):
                tag = "warn"; warn_cnt += 1
            elif iss:
                tag = "err"; err_cnt += 1
            else:
                tag = "ok"; ok_cnt += 1

            self.tree.insert("", "end",
                values=(ev, team, scores_str, total, issue_str, fname),
                tags=(tag,))

        # 명단 기준 미제출자
        if self.roster:
            for team, members in sorted(self.roster.items(),
                                        key=lambda x: int(x[0]) if x[0].isdigit() else x[0]):
                for mb in members:
                    if mb["display"] not in submitted:
                        self.tree.insert("", "end",
                            values=(mb["display"], team, "-", "-", "✘ 미제출", ""),
                            tags=("missing",))
                        missing_cnt += 1

        self.summary_var.set(
            f"제출 {len(self.records)}명  |  ✓ 정상 {ok_cnt}  |  "
            f"⚠ 경고 {warn_cnt}  |  ❌ 오류 {err_cnt}  |  ✘ 미제출 {missing_cnt}")

    def _on_double_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell": return
        col_id = self.tree.identify_column(event.x)
        if int(col_id.replace("#","")) - 1 != 1: return
        iid = self.tree.identify_row(event.y)
        if not iid: return
        row_idx = self.tree.index(iid)
        if row_idx >= len(self.records): return
        cur = self.records[row_idx].get("team","")
        new = self._ask_team(cur)
        if new is None: return
        self.records[row_idx]["team"] = new
        self.issues = detect_issues(self.records)
        self._populate()

    def _ask_team(self, current):
        dlg = tk.Toplevel(self)
        dlg.title("팀 번호 수정"); dlg.geometry("280x120")
        dlg.resizable(False,False); dlg.configure(bg="#F0F4F8"); dlg.grab_set()
        tk.Label(dlg, text="새 팀 번호:", font=("Malgun Gothic",10), bg="#F0F4F8").pack(pady=(16,4))
        var = tk.StringVar(value=current)
        entry = tk.Entry(dlg, textvariable=var, font=("Malgun Gothic",12), justify="center", width=10)
        entry.pack(); entry.select_range(0,"end"); entry.focus()
        result = [None]
        def ok(e=None): result[0]=var.get().strip(); dlg.destroy()
        entry.bind("<Return>", ok)
        bf2 = tk.Frame(dlg, bg="#F0F4F8"); bf2.pack(pady=8)
        tk.Button(bf2, text="확인", command=ok, bg=f"#{_NAVY}", fg="white",
                  font=("Malgun Gothic",9,"bold"), relief="flat", padx=12, pady=4).pack(side="left", padx=4)
        tk.Button(bf2, text="취소", command=dlg.destroy, bg="#7F8C8D", fg="white",
                  font=("Malgun Gothic",9,"bold"), relief="flat", padx=12, pady=4).pack(side="left", padx=4)
        dlg.wait_window()
        return result[0]

    def _generate(self):
        try:
            build_excel(self.records, self.out_path, self.roster)
            msg = f"✅  저장 완료!\n\n📁 {self.out_path}"
            if self.errors:
                msg += f"\n\n⚠️ 파싱 실패:\n" + "\n".join(self.errors)
            messagebox.showinfo("완료", msg, parent=self)
        except Exception as e:
            messagebox.showerror("오류", str(e), parent=self)


if __name__ == "__main__":
    App().mainloop()
